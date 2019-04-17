from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib import parse
import openpyxl
from selenium import webdriver

wd = openpyxl.load_workbook('search_rank_list_r.xlsx')
ws = wd.active
 
def bs(s_key, r_key):
    url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query="
    q = parse.quote(s_key)
    html = urlopen(url + q)
    bsObj = BeautifulSoup(html, "html.parser")
    dd = bsObj.find('dd','lst_relate')
    try:
        if r_key in dd.text:
            print(s_key)
            li = dd.find_all('li')
            for n, i in enumerate(li):
                if r_key == i.text.strip():
                    print('노출확인')
                    driver = webdriver.Chrome('./chromedriver.exe')
                    driver.get(url + q)
                    driver.save_screenshot(s_key+'.png')
                    driver.quit()
                    return str(n+1)
            else:
                print('없어요..')
                return None
    except AttributeError as e:
        print(e)
        return None

def relate_search():
    ws['a1'] = '순서'
    max_len = 0
    for r in ws.rows:
        if r[0].value == '순서':
            ws.cell(row=1, column=2).value = '키워드'
            ws.cell(row=1, column=3).value = '순위'
        else:
            row_index = r[0].row
            search_key = r[1].value
            relate_key = r[2].value
            # write excel
            ws.cell(row=row_index, column=1).value = row_index - 1
            ws.cell(row=row_index, column=2).value = search_key
            ws.cell(row=row_index, column=3).value = relate_key
            # start crawling
            print('start')
            print(search_key)
            ranking = bs(search_key,relate_key)
            if ranking == None:
                ws.cell(row=row_index, column=4).value = "x"
            else:
                ws.cell(row=row_index, column=4).value = ranking
    wd.save("result_search_rank_list.xlsx")
    wd.close()
 
relate_search()
