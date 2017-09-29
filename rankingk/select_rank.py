from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib import parse
import openpyxl
 
wd = openpyxl.load_workbook('select_rank_list.xlsx')
ws = wd.active
 
def bs(s_key, r_key):
    url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query="
    q = parse.quote(r_key)
    html = urlopen(url + q)
    bsObj = BeautifulSoup(html, "html.parser")
    dd = bsObj.find('dd','lst_relate')
    if bool(dd):
        li = dd.find_all('li')
        for n in range(len(li)):
            if li[n].text.find(str(s_key)) == 1:
                return n+1
                break
    else:
        return "연관검색어가 없습니다."

def relate_search():
    ws['a1'] = '순서'
    for r in ws.rows:
        if r[0].value == '순서':
            ws.cell(row=1, column=2).value = '키워드'
            ws.cell(row=1, column=3).value = '순위'
        else:
            for r in ws.rows:
                row_index = r[0].row
                search_key = r[1].value
                relate_key = r[2].value
                ranking = bs(search_key,relate_key)
                if ranking == None:
                    ranking = "x"
                ws.cell(row=row_index, column=1).value = row_index
                ws.cell(row=row_index, column=2).value = search_key
                ws.cell(row=row_index, column=3).value = relate_key
                ws.cell(row=row_index, column=4).value = ranking
                wd.save("result_select_ranking.xlsx")
    wd.close()
 
relate_search()
