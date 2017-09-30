from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib import parse
import openpyxl
 
wd = openpyxl.load_workbook('search_rank_list.xlsx')
ws = wd.active
 
def bs(s_key):
    url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query="
    q = parse.quote(s_key)
    html = urlopen(url + q)
    bsObj = BeautifulSoup(html, "html.parser")
    dd = bsObj.find('dd','lst_relate')
    list = []
    if bool(dd):
        li = dd.find_all('li')
        for n in range(len(li)):
            kwarg = li[n].text
            list.append(kwarg)
        return list
    else:
        return "순위가 없는 키워드입니다."

def relate_search():
    ws['a1'] = '순서'
    max_len = 0
    for r in ws.rows:
        if r[0].value == '순서':
            ws.cell(row=1, column=2).value = '키워드'
        else:
            row_index = r[0].row
            search_key = r[1].value
            ranking = bs(search_key)
            ws.cell(row=row_index, column=1).value = row_index - 1
            ws.cell(row=row_index, column=2).value = search_key
            if type(ranking) is str:
                ws.cell(row=row_index, column=3).value = ranking
            else:
                if len(ranking) > max_len:
                    max_len = len(ranking)
                for i in range(len(ranking)):
                    ws.cell(row=row_index, column=i+3).value = ranking[i]
    for i in range(max_len):
        ws.cell(row=1, column=i + 3).value = str(i+1) + '순위'
    wd.save("result_search_rank_list.xlsx")
    wd.close()
 
relate_search()
