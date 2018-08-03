import urllib.request
import urllib.parse
from datetime import datetime

from bs4 import BeautifulSoup as bs
from openpyxl import Workbook

client_id = 'xzbHzlxAEYYDo_DYeW6Y'
client_secret = 'l8ODdVqP_q'
key_list = []
with open('key_list.txt', 'rt') as file:
    for keyword in file:
        key_list.append(keyword[:-1])

url = 'https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query='

wb = Workbook()
ws = wb.active
ws['A1'] = 'time'
ws['B1'] = 'keyword'
ws['C1'] = 'rank'
ws['D1'] = 'title'
ws['E1'] = 'url'

def make_response(keyword):
    global url
    key_word = urllib.parse.quote(keyword)
    r_url = url + key_word
    request = urllib.request.Request(r_url)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header('X-Naver-Client-Secret', client_secret)
    response = urllib.request.urlopen(request)
    return response

def make_soup(response):
    soup = bs(response, 'html.parser')
    titles = soup.find_all('a', {'class':'lnk_tit'})
    urls = soup.find_all('a', {'class':'lnk_url'})
    return titles, urls

def write_excel(keyword, titles, urls, i):
    global ex_index
    now = datetime.now()
    ws['a'+str(ex_index)] = '%s-%s-%s %s-%s' % (now.year, now.month, now.day, now.hour, now.minute)
    ws['b'+str(ex_index)] = keyword
    for i in range(len(titles)):
        ws['c'+str(ex_index)] = str(i+1)+'위'
        ws['d'+str(ex_index)] = titles[i].text
        ws['e'+str(ex_index)] = urls[i].text
        ex_index += 1

ex_index = 2

for i, keyword in enumerate(key_list):
    print(str(i+1), ' ', keyword)
    response = make_response(keyword)
    titles, urls = make_soup(response)
    write_excel(keyword, titles, urls, i) 

wb.save('result_keyword_rank.xlsx')
