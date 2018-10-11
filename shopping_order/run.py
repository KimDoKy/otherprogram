# category : 품목. ex.설화사슴목장, 자연영어...
# path : 주문경로 ex. 위메프, 쇼핑몰..
# product : 상품 ex. 녹용, 사과식초..
# option : 상품 옵션 ex. 30포, 60포...
# count : 수량
# price : 결제금액
# delivery :택배비
# orderer : 주문자
# payee : 수령자
# call : 연락처
# address : 주소
# postal_code : 우편번호
# message : 배송메시지

import openpyxl
from openpyxl.styles import PatternFill, Font, Color
import os
from util import change_category, create_file
from market import eleven, shop, interpark, naverpay, gmarket, farm, wemape, wow_kim, imbak, tmon, coupang, kim

file_path = 'C:/Users/doky/Desktop/test/order'
os.chdir(file_path)
file_list = os.listdir()
order_file_list = [order for order in file_list if order[-5:] == '.xlsx' or order[-4:] == '.xls']

# 주문 파일 작업 시작(결과물 파일)
order_wb = openpyxl.Workbook()
order_ws = order_wb.create_sheet(index=0, title='filename')

order_ws['A1'] = '주문경로'
order_ws['B1'] = '품목'
order_ws['C1'] = '상품'
order_ws['D1'] = '옵션'
order_ws['E1'] = '수량'
order_ws['F1'] = '결제금액'
order_ws['G1'] = '택배비'
order_ws['H1'] = '주문자'
order_ws['I1'] = '수령자'
order_ws['J1'] = '연락처'
order_ws['K1'] = '주소'
order_ws['L1'] = '우편번호'
order_ws['M1'] = '배송메시지'

# 셀 고정
order_ws.freeze_panes = 'A2'

# 맨 윗줖 셀 글씨 굵기, 색 지정
colorFill = PatternFill(fgColor=Color('42c2e5'),
           fill_type='solid')

import string
alpha = string.ascii_lowercase[:13]

for bet in alpha:
    order_ws[bet + str(1)].fill = colorFill
    order_ws[bet + str(1)].font = Font(bold=True)


for file_name in order_file_list:
    if file_name[-5:] == '.xlsx':
        path = file_name[:-5]
    elif file_name[-4:] == '.xls':
        path = file_name[:-4]
    wb_str = "wb_%s = openpyxl.load_workbook('%s')" % (path, file_name)
    ws_str = "ws_%s = wb_%s.active" % (path, path)
    order_str = "order_list = ws_%s" % path
    close_str = "wb_%s.close()" % path
    print('--------------', path, '------------')
    # execute
    exec(wb_str)
    exec(ws_str)
    exec(order_str)

#     print("파일 로딩 성공 테스트")
#     for r in order_list.rows:
#         print(r[3].value)

    if path == '쇼핑몰':
        order_file, row_index = shop(order_list) 
    elif path == '티몬':
        order_file, row_index = tmon(order_list)
    elif path == '스토어팜':
        order_file, row_index = farm(order_list)
    elif path == 'eleven':
        order_file, row_index = eleven(order_list)
    elif path == '쿠팡':
        order_file, row_index = coupang(order_list)
    elif path == '인터파크':
        order_file, row_index = interpark(order_list)
    elif path == '네이버페이':
        order_file, row_index = naverpay(order_list)
    elif path == '지마켓' or path == '옥션':
        order_file, row_index = gmarket(order_list)
    elif path == '위메프':
        order_file, row_index = wemape(order_list)
    elif path == '김약사':
        order_file, row_index = kim(order_list)
    elif path == '와우김약사':
        order_file, row_index = wow_kim(order_list)
    elif path == '임박몰':
        order_file, row_index = imbak(order_list)
    else:
        order_file = {}
        
    exec(close_str)
    create_file(order_file, order_ws, row_index)

# 금일 날짜로 파일 이름 만들기
from datetime import datetime
result_file_name = datetime.today().strftime("%Y_%m_%d.xlsx")
print(result_file_name)

# 엑셀 파일 저장
order_wb.save(filename=result_file_name)

order_wb.close()