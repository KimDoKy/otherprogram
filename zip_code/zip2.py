from urllib.request import urlopen
from bs4 import BeautifulSoup
from urllib import parse
import openpyxl
 
wd = openpyxl.load_workbook('zip.xlsx')
ws = wd.active
 
def bs(adress):
    url = "https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query="
    q = parse.quote(adress)
    html = urlopen(url + q)
    bsObj = BeautifulSoup(html, "html.parser")
    span = bsObj.find("span", "zipcode")
    span2 = bsObj.find("p", "addr2")
    if bool(span):
        return span.text
    elif bool(span2):
        return span2.text
    else:
        info = "정보가 없습니다"
        return info
 
def zipcode():
    for r in ws.rows:
        row_index = r[0].row
        name = r[1].value
        phone = r[2].value
        adress = r[3].value
        zip_c = bs(adress)
        ws.cell(row=row_index, column=1).value = row_index
        ws.cell(row=row_index, column=2).value = name
        ws.cell(row=row_index, column=3).value = phone
        ws.cell(row=row_index, column=4).value = adress
        ws.cell(row=row_index, column=5).value = zip_c
        wd.save("zip_comp.xlsx")
    wd.close()
 
zipcode()
