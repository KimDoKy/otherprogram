import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill, Color


file_name = input(str('file_name: ')) + '.xlsx'
name = input(str('name: '))
print(file_name, name)
docu = openpyxl.load_workbook(file_name)
sheet = docu.get_sheet_by_name('call')

num = 0
# init total time
init_time = sheet['d2'].value
total_time = init_time - datetime.timedelta(
    days=init_time.day,
    hours=init_time.hour,
    minutes=init_time.minute,
    seconds=init_time.second) + datetime.timedelta(
	days=1)
print(total_time)

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = str(name)

for i, line in enumerate(sheet):
    name_1= line[6].value
    if name == name_1:
        num += 1
        print('call count : %i' % num)
        start = sheet['d'+ str(i+1)].value
        call_time = sheet['n'+ str(i+1)].value
        print(type(call_time))
        end = start + datetime.timedelta(
            minutes=call_time.minute,
            seconds=call_time.second)
    	# resulte insert excel
        ws1['a'+str(num+1)] = num
        ws1['b'+str(num+1)] = start
        ws1['c'+str(num+1)] = call_time
        ws1['d'+str(num+1)] = end
        total_time = total_time + datetime.timedelta(
            minutes=call_time.minute,
            seconds=call_time.second) + datetime.timedelta(days=1)
        print('after total: ' + str(total_time))
    	# 20minutes judge
        try:
            compare = ws1['d'+str(num+1)].value + datetime.timedelta(minutes=20)
            compare_2 = ws1['d'+str(num)].value
            if compare > compare_2 :
                ws1['e'+str(num+1)] = "good"
            else:
                ws1['e'+str(num+1)] = "bad"
                ws1['e'+str(num+1)].fill = PatternFill(
                    patternType='solid',
                    fgColor=Color('FFC000'))
            # settings excel head
            ws1['a1'] = 'no.'
            ws1['b1'] = 'start_call'
            ws1['c1'] = 'call_time'
            ws1['d1'] = 'end_call'
            ws1['e1'] = 'resulte'
            ws1['g1'] = 'call_resulte'
            ws1['g2'] = total_time
        except TypeError as e:
            print('error: ' + str(e))
            pass

fname = 'C:\\Users\doky\Desktop\call_profile\ ' + name + '.xlsx'
print(fname)
wb.save(fname)


